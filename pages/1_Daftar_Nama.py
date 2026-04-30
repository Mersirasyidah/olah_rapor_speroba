import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
import pytz
import os

# --- Konstanta Global ---
KEPALA_SEKOLAH = "Alina Fiftiyani Nurjannah, M.Pd."
NIP_KEPSEK = "19800105 200903 2 006"

# --- Daftar Mata Pelajaran Umum (untuk selectbox) ---
COMMON_SUBJECTS = [
    "Matematika", "Bahasa Indonesia", "IPA",
    "IPS", "Bahasa Inggris", "Pendidikan Agama Islam",
    "PP", "Seni Budaya",
    "PJOK", "Informatika",
    "Prakarya", "Bahasa Jawa"
]

# =========================================================
# Fungsi untuk membuat dokumen Excel "Daftar Siswa"
# =========================================================
def generate_excel_daftar_siswa(dataframe, kelas, semester, tahun_pelajaran, nama_wali_kelas, nip_wali_kelas):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Daftar Siswa {kelas}"

    # --- Pengaturan Halaman (Legal, Portrait) ---
    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL # Paper size Legal
    # Margins in inches (1 inch = 25.4 mm)
    ws.page_margins = PageMargins(left=28/25.4, right=18/25.4, top=24/25.4, bottom=19/25.4)
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT # Portrait
    ws.print_options.horizontalCentered = True # Rata Halaman Horizontal
    ws.print_options.verticalCentered = False # Rata Halaman Vertikal dinonaktifkan

    # --- Styles ---
    font_bold_12 = Font(name='Times New Roman', size=12, bold=True)
    font_normal_12 = Font(name='Times New Roman', size=12)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # --- Header Dokumen Excel ---
    current_row = 1
    # Merge cells for main title (spans 7 columns: A-G)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    cell_title = ws[f'A{current_row}']
    cell_title.value = "DAFTAR SISWA"
    cell_title.font = font_bold_12
    cell_title.alignment = center_align

    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    cell_subtitle = ws[f'A{current_row}']
    cell_subtitle.value = f"SMP NEGERI 2 BANGUNTAPAN TAHUN PELAJARAN {tahun_pelajaran}"
    cell_subtitle.font = font_bold_12
    cell_subtitle.alignment = center_align

    current_row += 2 # Spasi setelah judul

    # Information table
    # Kelas dan Wali Kelas di kolom A, isi kelas dan isi Nama Wali kelas di kolom D
    ws[f'A{current_row}'].value = "Kelas"
    ws[f'D{current_row}'].value = f": {kelas}"
    # Semester dan Tahun Pelajaran di kolom F, isi Semester dan isi Tahun Pelajaran di kolom G
    ws[f'F{current_row}'].value = "Semester"
    ws[f'G{current_row}'].value = f": {semester}"
    ws[f'F{current_row}'].alignment = Alignment(horizontal='left', vertical='center', indent=5)
    for col_letter in ['A', 'D', 'F', 'G']:
        ws[f'{col_letter}{current_row}'].font = font_normal_12
        ws[f'{col_letter}{current_row}'].alignment = left_align

    current_row += 1
    ws[f'A{current_row}'].value = "Wali Kelas" # Changed from "Nama Wali Kelas"
    ws[f'D{current_row}'].value = f": {nama_wali_kelas}"
    ws[f'F{current_row}'].value = "Tahun Pelajaran"
    ws[f'G{current_row}'].value = f": {tahun_pelajaran}"

    for col_letter in ['A', 'D', 'F', 'G']:
        ws[f'{col_letter}{current_row}'].font = font_normal_12
        ws[f'{col_letter}{current_row}'].alignment = left_align # All info cells left aligned and vertically centered

    current_row += 2 # Spasi sebelum tabel siswa

    # --- Tabel Siswa Header ---
    header_start_row = current_row

    # Define new column constants for the table based on swapped NIS and Nama Siswa
    COL_NO_START = 1
    COL_NO_END = 2 # Merged with COL_NO_START (A:B)
    COL_NIS_START = 3 # NIS is now column C
    COL_NIS_END = 4 # Merged with COL_NIS_START (C:D)
    COL_NAMA_START = 5 # Nama Siswa is now column E
    COL_NAMA_END = 6 # Merged with COL_NAMA_START (E:F)
    COL_JK = 7 # Jenis Kelamin is now column G
    TOTAL_TABLE_COLS = 7 # Total columns in this table

    # Merge headers for No.
    ws.merge_cells(start_row=header_start_row, start_column=COL_NO_START, end_row=header_start_row, end_column=COL_NO_END)
    ws.cell(row=header_start_row, column=COL_NO_START, value="No.").font = font_bold_12
    ws.cell(row=header_start_row, column=COL_NO_START).alignment = center_align

    # Merge headers for NIS
    ws.merge_cells(start_row=header_start_row, start_column=COL_NIS_START, end_row=header_start_row, end_column=COL_NIS_END)
    ws.cell(row=header_start_row, column=COL_NIS_START, value="NIS").font = font_bold_12
    ws.cell(row=header_start_row, column=COL_NIS_START).alignment = center_align

    # Merge headers for Nama Siswa
    ws.merge_cells(start_row=header_start_row, start_column=COL_NAMA_START, end_row=header_start_row, end_column=COL_NAMA_END)
    ws.cell(row=header_start_row, column=COL_NAMA_START, value="Nama Siswa").font = font_bold_12
    ws.cell(row=header_start_row, column=COL_NAMA_START).alignment = left_align # Nama Siswa header should be left aligned

    ws.cell(row=header_start_row, column=COL_JK, value="JK").font = font_bold_12
    ws.cell(row=header_start_row, column=COL_JK).alignment = center_align

    # Set header row height
    ws.row_dimensions[header_start_row].height = 21

    # Apply borders and vertical center alignment to header cells
    for c_idx in range(1, TOTAL_TABLE_COLS + 1):
        cell = ws.cell(row=header_start_row, column=c_idx)
        cell.border = thin_border
        # Alignment already set for each cell above, no need to re-apply

    current_row += 1 # Pindah ke baris untuk entri data

    # --- Isi Data Siswa ---
    for i, row_data in enumerate(dataframe.itertuples(), 1):
        data_row_idx = current_row + i - 1

        # Merge cells for No. and NIS for each data row
        ws.merge_cells(start_row=data_row_idx, start_column=COL_NO_START, end_row=data_row_idx, end_column=COL_NO_END)
        ws.merge_cells(start_row=data_row_idx, start_column=COL_NIS_START, end_row=data_row_idx, end_column=COL_NIS_END)
        # Merge cells for Nama Siswa for each data row
        ws.merge_cells(start_row=data_row_idx, start_column=COL_NAMA_START, end_row=data_row_idx, end_column=COL_NAMA_END)

        ws.cell(row=data_row_idx, column=COL_NO_START, value=i).font = font_normal_12
        ws.cell(row=data_row_idx, column=COL_NIS_START, value=str(getattr(row_data, "NIS"))).font = font_normal_12
        ws.cell(row=data_row_idx, column=COL_NAMA_START, value=getattr(row_data, "Nama")).font = font_normal_12
        ws.cell(row=data_row_idx, column=COL_JK, value=getattr(row_data, "Jenis_Kelamin")).font = font_normal_12

        # Set row height for data rows
        ws.row_dimensions[data_row_idx].height = 17.25

        # Apply borders and alignment to data cells
        for col_idx in range(1, TOTAL_TABLE_COLS + 1):
            cell = ws.cell(row=data_row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx == COL_NAMA_START: # Nama Siswa
                cell.alignment = left_align
            else:
                cell.alignment = center_align

    # Atur lebar kolom
    ws.column_dimensions[get_column_letter(COL_NO_START)].width = 2.29 # No (merged)
    ws.column_dimensions[get_column_letter(COL_NO_END)].width = 2.14 # Hide the merged part
    ws.column_dimensions[get_column_letter(COL_NIS_START)].width = 6.14 # NIS (merged)
    ws.column_dimensions[get_column_letter(COL_NIS_END)].width = 1.71 # Hide the merged part
    ws.column_dimensions[get_column_letter(COL_NAMA_START)].width = 22.71 # Nama Siswa (merged)
    ws.column_dimensions[get_column_letter(COL_NAMA_END)].width = 27 # Hide the merged part
    ws.column_dimensions[get_column_letter(COL_JK)].width = 7.86 # Jenis Kelamin

    current_row += len(dataframe) + 1 # Pindah ke bawah tabel, 1 baris spasi

    # --- Bagian Keterangan ---
    if 'Jenis_Kelamin' in dataframe.columns:
        jumlah_L = dataframe[dataframe['Jenis_Kelamin'].astype(str).str.upper() == 'L'].shape[0]
        jumlah_P = dataframe[dataframe['Jenis_Kelamin'].astype(str).str.upper() == 'P'].shape[0]

        ws.cell(row=current_row, column=1, value="Keterangan:").font = font_bold_12
        ws.cell(row=current_row, column=1).alignment = left_align
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"- Jumlah Laki-laki (L) : {jumlah_L}").font = font_normal_12
        ws.cell(row=current_row, column=1).alignment = left_align
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"- Jumlah Perempuan (P) : {jumlah_P}").font = font_normal_12
        ws.cell(row=current_row, column=1).alignment = left_align
        current_row += 1

    current_row += 1 # Tambah 1 baris kosong antara Keterangan dan Mengetahui

    # --- Bagian Tanda Tangan ---
    #Mengetahui (Kepala Sekolah) di kolom A
    ws.cell(row=current_row, column=1, value="Mengetahui").font = font_normal_12
    ws.cell(row=current_row, column=1).alignment = left_align
    ws.cell(row=current_row + 1, column=1, value="Kepala Sekolah").font = font_normal_12
    ws.cell(row=current_row + 1, column=1).alignment = left_align

    #Wali Kelas dan Bantul di kolom F (index 7)
    ws.cell(row=current_row, column=6, value=f"Bantul, ............................... {tahun_pelajaran.split('/')[0]}").font = font_normal_12
    ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='left', indent=7)
    #ws.cell(row=current_row, column=6).alignment = left_align
    ws.cell(row=current_row + 1, column=6, value=f"Wali Kelas {kelas}").font = font_normal_12
    ws.cell(row=current_row + 1, column=6).alignment = Alignment(horizontal='left', indent=7)
    #ws.cell(row=current_row + 1, column=6).alignment = left_align
    current_row += 2

    current_row += 3 # Jarak untuk tanda tangan (3 baris)
    ws.cell(row=current_row, column=1, value=KEPALA_SEKOLAH).font = font_normal_12
    ws.cell(row=current_row, column=1).alignment = left_align
    ws.cell(row=current_row, column=6, value=nama_wali_kelas).font = font_normal_12
    ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='left', indent=7)

    current_row += 1 # Pindah ke baris berikutnya untuk NIP
    ws.cell(row=current_row, column=1, value=f"NIP. {NIP_KEPSEK}").font = font_normal_12
    ws.cell(row=current_row, column=1).alignment = left_align
    ws.cell(row=current_row, column=6, value=f"NIP. {nip_wali_kelas}").font = font_normal_12
    ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='left', indent=7)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ===================== APLIKASI STREAMLIT UTAMA =====================
st.set_page_config(layout="wide", page_title="Aplikasi Manajemen Data Siswa")
st.title("Aplikasi Manajemen Data Siswa")

default_csv_path = "daftar_siswa.csv"
df = None

# --- Memuat File CSV ---
# Menggunakan st.session_state untuk menjaga df setelah diunggah/dimuat
if 'df_loaded' not in st.session_state:
    st.session_state.df_loaded = None

if st.session_state.df_loaded is None:
    if os.path.exists(default_csv_path):
        try:
            df = pd.read_csv(default_csv_path)
            st.session_state.df_loaded = df
            #st.success(f"File '{default_csv_path}' berhasil dimuat secara otomatis.")
        except Exception as e:
            st.error(f"Gagal memuat '{default_csv_path}': {e}. Coba unggah file manual.")

    if st.session_state.df_loaded is None: # Jika belum dimuat atau ada error
        st.warning(f"File '{default_csv_path}' tidak ditemukan atau gagal dimuat. Harap unggah file siswa.")
        uploaded_file_obj = st.file_uploader("Unggah file CSV/Excel daftar siswa (Harus punya kolom 'Kelas', 'Nama', 'NIS', 'Jenis_Kelamin')", type=["csv", "xlsx"], key="manual_upload")
        if uploaded_file_obj is not None:
            try:
                if uploaded_file_obj.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file_obj)
                else:
                    df = pd.read_excel(uploaded_file_obj)
                st.session_state.df_loaded = df
            except Exception as e:
                st.error(f"Terjadi kesalahan saat membaca file: {e}. Pastikan format file benar.")
                st.session_state.df_loaded = None
else:
    df = st.session_state.df_loaded # Ambil df dari session_state jika sudah dimuat

# Hentikan aplikasi jika DataFrame belum dimuat
if df is None:
    st.info("Silakan unggah file siswa untuk melanjutkan.")
    st.stop()

# Periksa kolom-kolom penting
required_cols_ds = ['Kelas', 'Nama', 'NIS', 'Jenis_Kelamin']
missing_cols_ds = [col for col in required_cols_ds if col not in df.columns]

if missing_cols_ds:
    st.error(f"Kolom yang diperlukan tidak ditemukan di file siswa Anda untuk 'Daftar Siswa': {', '.join(missing_cols_ds)}. Harap perbaiki file Anda.")
    st.stop()

st.markdown("---") # Garis pemisah


# Opsi kelas yang tersedia
kelas_options = sorted(df['Kelas'].unique())
if not kelas_options:
    st.error("Tidak ada data kelas yang ditemukan di kolom 'Kelas' file siswa Anda.")
    st.stop()

# --- Bagian 1: Cetak Daftar Siswa ---
st.header("📋 Cetak Daftar Siswa")

kelas_daftar_siswa = st.selectbox(
    "Pilih Kelas untuk Daftar Siswa",
    kelas_options,
    key="kelas_daftar_siswa_selectbox"
)
data_kelas_daftar_siswa = df[df['Kelas'] == kelas_daftar_siswa].reset_index(drop=True)

st.dataframe(data_kelas_daftar_siswa)

# Tambahkan informasi jumlah Laki-laki dan Perempuan di Streamlit
if 'Jenis_Kelamin' in data_kelas_daftar_siswa.columns:
    jumlah_L_selected_class = data_kelas_daftar_siswa[data_kelas_daftar_siswa['Jenis_Kelamin'].astype(str).str.upper() == 'L'].shape[0]
    jumlah_P_selected_class = data_kelas_daftar_siswa[data_kelas_daftar_siswa['Jenis_Kelamin'].astype(str).str.upper() == 'P'].shape[0]
    st.write(f"**Keterangan:**")
    st.write(f"- Jumlah Laki-laki (L) : {jumlah_L_selected_class}")
    st.write(f"- Jumlah Perempuan (P) : {jumlah_P_selected_class}")

nama_wali_kelas = st.text_input("Nama Wali Kelas", key="nama_wali_kelas_ds")
nip_wali_kelas = st.text_input("NIP Wali Kelas", key="nip_wali_kelas_ds")
semester_ds = st.selectbox(
    "Semester (Daftar Siswa)",
    ["Ganjil", "Genap"],
    key="semester_ds_selectbox"
)
#current_year = datetime.now().year
#tahun_options = [f"{y}/{y+1}" for y in range(current_year + 1, current_year - 5, -1)]
start_year = 2025
tahun_options = [f"{y}/{y+1}" for y in range(start_year, start_year + 5)] # Contoh: 2025/2026 hingga 2029/2030
tahun_ds = st.selectbox(
    "Tahun Pelajaran (Daftar Siswa)",
    tahun_options,
    index=0,
    key="tahun_ds_selectbox"
)

# Mengubah tombol generate dan download untuk Excel
if st.button("🔽 Generate Daftar Siswa (Excel)", key="generate_ds_button"):
    if not nama_wali_kelas or not nip_wali_kelas:
        st.error("Nama Wali Kelas dan NIP Wali Kelas tidak boleh kosong.")
    else:
        excel_buffer_ds = generate_excel_daftar_siswa(data_kelas_daftar_siswa, kelas_daftar_siswa, semester_ds, tahun_ds, nama_wali_kelas, nip_wali_kelas)
        st.download_button(
            label="Unduh Daftar Siswa (Excel)",
            data=excel_buffer_ds,
            file_name=f"Daftar_Siswa_{kelas_daftar_siswa}_{tahun_ds}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_ds_button"
        )
        st.success("Daftar Siswa berhasil dibuat!")

st.markdown("---")
st.write(f"Created by Mersi")
