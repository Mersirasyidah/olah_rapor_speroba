import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
import pytz
import os

# --- Konstanta Global ---
KEPALA_SEKOLAH = "Alina Fiftiyani Nurjannah, M.Pd."
NIP_KEPSEK = "19800105 200903 2 006"
FILE_SISWA_DEFAULT = "daftar_siswa.csv"

# --- Daftar Pilihan untuk Filter ---
COMMON_SUBJECTS = [
    "Matematika", "Bahasa Indonesia", "IPA",
    "IPS", "Bahasa Inggris", "Pendidikan Agama Islam",
    "PP", "Seni Budaya",
    "PJOK", "Informatika",
    "Prakarya", "Bahasa Jawa"
]

# Generate daftar kelas (7A-9E)
CLASSES = []
for grade in [7, 8, 9]:
    for letter in ['A', 'B', 'C', 'D', 'E']:
        CLASSES.append(f"{grade}{letter}")

# Generate daftar tahun pelajaran (mulai 2025/2026)
current_year = datetime.now().year
start_year = 2025
if current_year > start_year:
    start_year = current_year

if start_year % 2 != 0:
    start_year -= 1

YEAR_OPTIONS = []
for year in range(start_year, start_year + 5):
    YEAR_OPTIONS.append(f"{year}/{year+1}")


# =========================================================
# FUNGSI EXCEL (TIDAK ADA PERUBAHAN DARI KODE ASLI)
# =========================================================
def generate_excel_absensi_panjang(dataframe, mapel, semester, kelas, tahun_pelajaran, guru, nip_guru):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Daftar Hadir {kelas}"

    # --- Pengaturan Halaman (Legal, Portrait) ---
    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
    ws.page_margins = PageMargins(
        left=18/25.4,
        right=0.8/25.4,
        top=19/25.4,
        bottom=24/25.4
    )
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    # --- Styles ---
    font_bold_12 = Font(name='Times New Roman', size=12, bold=True)
    font_normal_12 = Font(name='Times New Roman', size=12)
    font_bold_9 = Font(name='Times New Roman', size=9, bold=True)
    font_normal_7 = Font(name='Times New Roman', size=7)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # --- Header Informasi (Baris atas) ---
    current_row = 1

    COL_NO_START = 1
    COL_NO_END = 2
    COL_NAMA_START = 3
    COL_NAMA_END = 4
    COL_PERTEMUAN_START = 5
    COL_PERTEMUAN_END = 24
    COL_JUMLAH_START = 25
    COL_JUMLAH_END = 27
    TOTAL_EFFECTIVE_COLS = 27

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=TOTAL_EFFECTIVE_COLS)
    cell_title = ws[f'A{current_row}']
    cell_title.value = "DAFTAR HADIR SISWA"
    cell_title.font = font_bold_12
    cell_title.alignment = center_align

    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=TOTAL_EFFECTIVE_COLS)
    cell_subtitle = ws[f'A{current_row}']
    cell_subtitle.value = f"SMP NEGERI 2 BANGUNTAPAN TAHUN PELAJARAN {tahun_pelajaran}"
    cell_subtitle.font = font_bold_12
    cell_subtitle.alignment = center_align

    current_row += 2

    # Information table (Mata Pelajaran, Kelas, Semester, Tahun Pelajaran)
    ws[f'A{current_row}'].value = "Mata Pelajaran"
    ws[f'D{current_row}'].value = f": {mapel}"
    ws[f'K{current_row}'].value = "Semester"
    ws[f'T{current_row}'].value = f": {semester}"

    for col_letter in ['A', 'D', 'K', 'T']:
        ws[f'{col_letter}{current_row}'].font = font_normal_12
        ws[f'{col_letter}{current_row}'].alignment = left_align

    current_row += 1
    ws[f'A{current_row}'].value = "Kelas"
    ws[f'D{current_row}'].value = f": {kelas}"
    ws[f'K{current_row}'].value = "Tahun Pelajaran"
    ws[f'T{current_row}'].value = f": {tahun_pelajaran}"

    for col_letter in ['A', 'D', 'K', 'T']:
        ws[f'{col_letter}{current_row}'].font = font_normal_12
        ws[f'{col_letter}{current_row}'].alignment = left_align

    current_row += 2

    # --- Tabel Utama Header ---
    header_start_row = current_row

    # --- Header "No." ---
    ws.merge_cells(start_row=header_start_row, start_column=COL_NO_START, end_row=header_start_row + 2, end_column=COL_NO_END)
    cell_no_header = ws.cell(row=header_start_row, column=COL_NO_START)
    cell_no_header.value = "No."
    cell_no_header.font = font_bold_12
    cell_no_header.alignment = center_align

    # --- Header "Nama" ---
    ws.merge_cells(start_row=header_start_row, start_column=COL_NAMA_START, end_row=header_start_row + 2, end_column=COL_NAMA_END)
    cell_nama_header = ws.cell(row=header_start_row, column=COL_NAMA_START)
    cell_nama_header.value = "Nama"
    cell_nama_header.font = font_bold_12
    cell_nama_header.alignment = center_align

    # --- Header "Pertemuan ke..... Tanggal..." ---
    ws.merge_cells(start_row=header_start_row, start_column=COL_PERTEMUAN_START, end_row=header_start_row, end_column=COL_PERTEMUAN_END)
    cell_pertemuan_header = ws.cell(row=header_start_row, column=COL_PERTEMUAN_START)
    cell_pertemuan_header.value = "Pertemuan ke... Tanggal..."
    cell_pertemuan_header.font = font_bold_12
    cell_pertemuan_header.alignment = center_align

    # --- Header "Jumlah" ---
    ws.merge_cells(start_row=header_start_row, start_column=COL_JUMLAH_START, end_row=header_start_row, end_column=COL_JUMLAH_END)
    cell_jumlah_header = ws.cell(row=header_start_row, column=COL_JUMLAH_START)
    cell_jumlah_header.value = "Jumlah"
    cell_jumlah_header.font = font_bold_9
    cell_jumlah_header.alignment = center_align

    # --- Sub-headers (Row 2 and 3 relative to header_start_row) ---
    for i in range(20):
        col_idx = COL_PERTEMUAN_START + i
        cell = ws.cell(row=header_start_row + 1, column=col_idx)
        cell.value = str(i + 1)
        cell.font = font_normal_7
        cell.alignment = center_align

        cell_empty = ws.cell(row=header_start_row + 2, column=col_idx)
        cell_empty.value = ""
        cell_empty.font = font_normal_7
        cell_empty.alignment = center_align

    # A, I, S headers
    ws.merge_cells(start_row=header_start_row + 1, start_column=COL_JUMLAH_START, end_row=header_start_row + 2, end_column=COL_JUMLAH_START)
    cell_a_header = ws.cell(row=header_start_row + 1, column=COL_JUMLAH_START)
    cell_a_header.value = "A"
    cell_a_header.font = font_bold_9
    cell_a_header.alignment = center_align

    ws.merge_cells(start_row=header_start_row + 1, start_column=COL_JUMLAH_START + 1, end_row=header_start_row + 2, end_column=COL_JUMLAH_START + 1)
    cell_i_header = ws.cell(row=header_start_row + 1, column=COL_JUMLAH_START + 1)
    cell_i_header.value = "I"
    cell_i_header.font = font_bold_9
    cell_i_header.alignment = center_align

    ws.merge_cells(start_row=header_start_row + 1, start_column=COL_JUMLAH_START + 2, end_row=header_start_row + 2, end_column=COL_JUMLAH_START + 2)
    cell_s_header = ws.cell(row=header_start_row + 1, column=COL_JUMLAH_START + 2)
    cell_s_header.value = "S"
    cell_s_header.font = font_bold_9
    cell_s_header.alignment = center_align

    # Apply borders
    for r_idx in range(header_start_row, header_start_row + 3):
        for c_idx in range(1, TOTAL_EFFECTIVE_COLS + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = thin_border

    # Set dimensions
    ws.row_dimensions[header_start_row].height = 20
    ws.row_dimensions[header_start_row + 1].height = 20
    ws.row_dimensions[header_start_row + 2].height = 20

    ws.column_dimensions[get_column_letter(COL_NO_START)].width = 2.57
    ws.column_dimensions[get_column_letter(COL_NO_END)].width = 3.43
    ws.column_dimensions[get_column_letter(COL_NAMA_START)].width = 8.14
    ws.column_dimensions[get_column_letter(COL_NAMA_END)].width = 26.86
    for i in range(COL_PERTEMUAN_START, COL_PERTEMUAN_END + 1):
        ws.column_dimensions[get_column_letter(i)].width = 2.3
    for i in range(COL_JUMLAH_START, COL_JUMLAH_END + 1):
        ws.column_dimensions[get_column_letter(i)].width = 2.7

    current_row += 3

    # --- Isi Data Siswa (yang sudah difilter) ---
    for i, row_data in enumerate(dataframe.itertuples(), 1):
        data_row_idx = current_row + i - 1

        ws.merge_cells(start_row=data_row_idx, start_column=COL_NO_START, end_row=data_row_idx, end_column=COL_NO_END)
        ws.merge_cells(start_row=data_row_idx, start_column=COL_NAMA_START, end_row=data_row_idx, end_column=COL_NAMA_END)

        ws.cell(row=data_row_idx, column=COL_NO_START, value=i).font = font_normal_12
        ws.cell(row=data_row_idx, column=COL_NO_START).alignment = center_align

        ws.cell(row=data_row_idx, column=COL_NAMA_START, value=getattr(row_data, "Nama")).font = font_normal_12
        ws.cell(row=data_row_idx, column=COL_NAMA_START).alignment = left_align

        # Apply border
        for c_idx in range(1, TOTAL_EFFECTIVE_COLS + 1):
            cell = ws.cell(row=data_row_idx, column=c_idx)
            cell.border = thin_border

        ws.row_dimensions[data_row_idx].height = 20

    current_row += len(dataframe) + 1

    # --- Bagian Tanda Tangan ---
    ws.cell(row=current_row, column=3, value="Mengetahui").font = font_normal_12
    ws.cell(row=current_row, column=3).alignment = left_align
    current_row += 1
    ws.cell(row=current_row, column=3, value="Kepala Sekolah").font = font_normal_12
    ws.cell(row=current_row, column=3).alignment = left_align

    tz_jakarta = pytz.timezone('Asia/Jakarta')
    # Menggunakan tahun yang dipilih untuk tanggal Bantul jika lebih masuk akal, tapi karena ini adalah file template,
    # menggunakan tahun saat ini (now.year) sudah cukup.
    now = datetime.now(tz_jakarta)
    tanggal_formatted = f"Bantul, .................................... {now.year}"

    ws.cell(row=current_row - 1, column=12, value=tanggal_formatted).font = font_normal_12
    ws.cell(row=current_row - 1, column=12).alignment = left_align
    ws.cell(row=current_row, column=12, value=f"Guru {mapel}").font = font_normal_12
    ws.cell(row=current_row, column=12).alignment = left_align

    current_row += 3

    ws.cell(row=current_row, column=3, value=KEPALA_SEKOLAH).font = font_normal_12
    ws.cell(row=current_row, column=3).alignment = left_align
    ws.cell(row=current_row, column=12, value=guru).font = font_normal_12
    ws.cell(row=current_row, column=12).alignment = left_align

    current_row += 1
    ws.cell(row=current_row, column=3, value=f"NIP. {NIP_KEPSEK}").font = font_normal_12
    ws.cell(row=current_row, column=3).alignment = left_align
    ws.cell(row=current_row, column=12, value=f"NIP. {nip_guru}").font = font_normal_12
    ws.cell(row=current_row, column=12).alignment = left_align

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =========================================================
# LOGIKA UTAMA STREAMLIT (DENGAN PEMFILTERAN KELAS)
# =========================================================

st.set_page_config(page_title="Generator Daftar Hadir Siswa", layout="wide")

st.title("🗃️ Generator Daftar Hadir Siswa")

# --- Muat File Default / Upload ---
df_siswa_global = None
file_source_msg = ""
data_loaded_successfully = False

if os.path.exists(FILE_SISWA_DEFAULT):
    try:
        # Pemuatan file default
        df_siswa_global = pd.read_csv(FILE_SISWA_DEFAULT)
        if 'Nama' in df_siswa_global.columns:
            #file_source_msg = f"Berhasil memuat file **{FILE_SISWA_DEFAULT}** ({len(df_siswa_global)} siswa)."
            data_loaded_successfully = True
        else:
            file_source_msg = f"⚠️ File **{FILE_SISWA_DEFAULT}** tidak memiliki kolom 'Nama'. Silakan unggah file secara manual."
    except Exception as e:
        file_source_msg = f"⚠️ Gagal memuat {FILE_SISWA_DEFAULT}. Silakan unggah file secara manual. Error: {e}"

if not data_loaded_successfully:
    uploaded_file = st.file_uploader(f"Unggah File Excel/CSV Daftar Siswa (Harus memiliki kolom 'Nama' dan **'Kelas'**). {file_source_msg}", type=["xlsx", "xls", "csv"])
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith('.csv'):
                df_siswa_global = pd.read_csv(uploaded_file)
            else:
                df_siswa_global = pd.read_excel(uploaded_file)

            if 'Nama' in df_siswa_global.columns:
                data_loaded_successfully = True
                st.success(f"File berhasil diunggah. Ditemukan {len(df_siswa_global)} siswa.")
            else:
                st.error("File siswa harus memiliki kolom **'Nama'**.")
                df_siswa_global = None

        except Exception as e:
            st.error(f"Terjadi kesalahan saat memproses file yang diunggah: {e}")
else:
    st.info(file_source_msg)


# --- Input Metadata & Pemfilteran ---
if data_loaded_successfully and df_siswa_global is not None:

    # Cek Kolom Kelas
    if 'Kelas' not in df_siswa_global.columns:
        st.error("🚨 **ERROR:** Database siswa tidak memiliki kolom **'Kelas'**! Pemfilteran per kelas tidak dapat dilakukan.")
        st.dataframe(df_siswa_global.head(), use_container_width=True)
        # Hentikan proses lebih lanjut
        data_loaded_successfully = False

if data_loaded_successfully and df_siswa_global is not None:
    st.dataframe(df_siswa_global[['Nama', 'Kelas']].head(), use_container_width=True) # Tampilkan Kolom Kelas

    st.subheader("Data Kelas dan Guru")
    col1, col2, col3 = st.columns(3)

    with col1:
        selected_mapel = st.selectbox("Mata Pelajaran:", COMMON_SUBJECTS)
        # FILTER KELAS BARU
        selected_kelas = st.selectbox("Kelas yang Akan Digenerate:", CLASSES, index=CLASSES.index("7 A") if "7 A" in CLASSES else 0)

    with col2:
        # FILTER TAHUN PELAJARAN BARU
        selected_tahun_pelajaran = st.selectbox("Tahun Pelajaran:", YEAR_OPTIONS)
        guru = st.text_input("Nama Guru Mata Pelajaran", "")

    with col3:
        semester = st.selectbox("Semester:", ["Ganjil", "Genap"])
        nip_guru = st.text_input("NIP Guru Mata Pelajaran", "")

    st.divider()

    if st.button("Generate File Daftar Hadir Excel 📊"):
        if guru and nip_guru and selected_kelas and selected_tahun_pelajaran:

            # --- PEMFILTERAN DATA SESUAI KELAS YANG DIPILIH ---
            df_filtered = df_siswa_global.query('Kelas == @selected_kelas')

            if df_filtered.empty:
                st.warning(f"Tidak ada siswa yang ditemukan di kelas **{selected_kelas}**. Cek kembali data di file siswa Anda.")
            else:
                st.info(f"Memproses {len(df_filtered)} siswa untuk kelas **{selected_kelas}**.")
                try:
                    excel_file_buffer = generate_excel_absensi_panjang(
                        df_filtered, # GUNAKAN DATAFRAME YANG SUDAH DIFILTER
                        selected_mapel,
                        semester,
                        selected_kelas,
                        selected_tahun_pelajaran,
                        guru,
                        nip_guru
                    )

                    file_name = f"Daftar Hadir Siswa {selected_mapel} {selected_kelas.replace(' ', '_')} Sem {semester} TP {selected_tahun_pelajaran.replace('/', '-')}.xlsx"

                    st.download_button(
                        label=f"Unduh Daftar Hadir Kelas {selected_kelas}",
                        data=excel_file_buffer,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.balloons()
                    st.success("File Daftar Hadir Siswa siap diunduh.")
                except Exception as e:
                    st.error(f"Terjadi kesalahan saat membuat Excel: {e}")
        else:
            st.warning("Mohon lengkapi semua data Kelas, Tahun Pelajaran, dan Guru di atas sebelum meng-generate.")

elif not data_loaded_successfully:
    st.error("Aplikasi tidak dapat berjalan tanpa data siswa yang valid.")
