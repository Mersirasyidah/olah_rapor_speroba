import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
import pytz
import os

# --- Konstanta Global ---
KEPALA_SEKOLAH = "Alina Fiftiyani Nurjannah, M.Pd."
NIP_KEPSEK = "19800105 200903 2 006"

# --- Daftar Mata Pelajaran Umum (untuk selectbox) ---
COMMON_SUBJECTS = [
    "Matematika", "Bahasa Indonesia", "IPA",
    "IPS", "Bahasa Inggris", "Pendidikan Agama Islam",
    "PP", "Seni Budaya",
    "PJOK", "Informatika",
    "Prakarya", "Bahasa Jawa"
]

# --- Variabel Database Default ---
default_csv_path = "daftar_siswa.csv"
df = None

# =========================================================
# Fungsi Utility untuk Opsi Kelas dan Tahun Pelajaran
# =========================================================

def generate_class_options():
    """Menghasilkan daftar opsi kelas dari 7A hingga 9E (Format Tanpa Spasi)."""
    classes = []
    # PERBAIKAN: Menghilangkan spasi saat menggabungkan level dan section.
    for level in ['7', '8', '9']:
        for section in ['A', 'B', 'C', 'D', 'E']:
            classes.append(f"{level}{section}") # Output: '7A', '7B', dst.
    return classes

def generate_academic_year_options(start_year=2025):
    """Menghasilkan daftar Tahun Pelajaran mulai dari 2025/2026 ke atas."""
    years = []
    current_year = datetime.now().year

    actual_start_year = max(start_year, current_year)
    limit_year = current_year + 5

    for year in range(actual_start_year, limit_year):
        years.append(f"{year}/{year+1}")

    if '2025/2026' not in years:
        years.insert(0, '2025/2026')

    years = sorted(list(set(years)), key=lambda x: int(x.split('/')[0]))
    return years


# =========================================================
# Fungsi untuk membuat dokumen Excel "Form Nilai Siswa"
# =========================================================
def generate_excel_form_nilai_siswa(dataframe, mapel, semester, kelas, tahun_pelajaran, guru, nip_guru):
    wb = Workbook()
    ws = wb.active
    # Menggunakan kelas tanpa spasi untuk judul sheet
    ws.title = f"Form Nilai {kelas}"

    # --- Pengaturan Halaman (Legal, Portrait) ---
    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
    ws.page_margins = PageMargins(
        left=1.8/2.54, right=1.8/2.54, top=1.9/2.54, bottom=1.9/2.54
    )
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    # --- Styles ---
    font_bold_12 = Font(name='Times New Roman', size=12, bold=True)
    font_normal_12 = Font(name='Times New Roman', size=12)
    font_bold_9 = Font(name='Times New Roman', size=9, bold=True)
    font_normal_9 = Font(name='Times New Roman', size=9)

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # --- Header Informasi (Baris atas) ---
    current_row = 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=20)
    cell_title = ws[f'A{current_row}']
    cell_title.value = "FORM NILAI SISWA"
    cell_title.font = font_bold_12
    cell_title.alignment = center_align

    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=20)
    cell_subtitle = ws[f'A{current_row}']
    cell_subtitle.value = f"SMP NEGERI 2 BANGUNTAPAN TAHUN PELAJARAN {tahun_pelajaran}"
    cell_subtitle.font = font_bold_12
    cell_subtitle.alignment = center_align

    current_row += 2

    # Information table
    ws[f'A{current_row}'].value = "Mata Pelajaran"
    ws[f'G{current_row}'].value = f": {mapel}"
    ws[f'J{current_row}'].value = "Kelas"
    ws[f'N{current_row}'].value = f": {kelas}" # Kelas tanpa spasi

    for col in ['A', 'G', 'J', 'N']:
        ws[f'{col}{current_row}'].font = font_normal_12
        ws[f'{col}{current_row}'].alignment = left_align

    current_row += 1
    ws[f'A{current_row}'].value = "Semester"
    ws[f'G{current_row}'].value = f": {semester}"
    ws[f'J{current_row}'].value = "Nama Guru"
    ws[f'N{current_row}'].value = f": {guru}"

    for col in ['A', 'G', 'J', 'N']:
        ws[f'{col}{current_row}'].font = font_normal_12
        ws[f'{col}{current_row}'].alignment = left_align

    current_row += 2

    # --- Tabel Utama Header ---
    header_start_row = current_row

    # Column Constants for Form Nilai Siswa
    COL_NO_START = 1; COL_NO_END = 2
    COL_NIS_START = 3; COL_NIS_END = 5
    COL_NAMA_START = 6; COL_NAMA_END = 7
    COL_FORMATIF_START = 8; COL_FORMATIF_END = 12
    COL_SUMATIF_START = 13; COL_SUMATIF_END = 17
    COL_PTS = 18
    COL_SAS_SAT = 19
    COL_NR = 20
    TOTAL_EFFECTIVE_COLS_NILAI = COL_NR

    # Row 1 Headers (Merged across 3 rows)
    ws.merge_cells(start_row=header_start_row, start_column=COL_NO_START, end_row=header_start_row + 2, end_column=COL_NO_END)
    ws.cell(row=header_start_row, column=COL_NO_START, value="No.").font = font_bold_12; ws.cell(row=header_start_row, column=COL_NO_START).alignment = center_align
    ws.merge_cells(start_row=header_start_row, start_column=COL_NIS_START, end_row=header_start_row + 2, end_column=COL_NIS_END)
    ws.cell(row=header_start_row, column=COL_NIS_START, value="NIS").font = font_bold_12; ws.cell(row=header_start_row, column=COL_NIS_START).alignment = center_align
    ws.merge_cells(start_row=header_start_row, start_column=COL_NAMA_START, end_row=header_start_row + 2, end_column=COL_NAMA_END)
    ws.cell(row=header_start_row, column=COL_NAMA_START, value="Nama Siswa").font = font_bold_12; ws.cell(row=header_start_row, column=COL_NAMA_START).alignment = center_align
    ws.merge_cells(start_row=header_start_row, start_column=COL_PTS, end_row=header_start_row + 2, end_column=COL_PTS)
    ws.cell(row=header_start_row, column=COL_PTS, value="PTS").font = font_bold_9; ws.cell(row=header_start_row, column=COL_PTS).alignment = center_align
    ws.merge_cells(start_row=header_start_row, start_column=COL_SAS_SAT, end_row=header_start_row + 2, end_column=COL_SAS_SAT)
    ws.cell(row=header_start_row, column=COL_SAS_SAT, value="SAS/SAT").font = font_bold_9; ws.cell(row=header_start_row, column=COL_SAS_SAT).alignment = center_align
    ws.merge_cells(start_row=header_start_row, start_column=COL_NR, end_row=header_start_row + 2, end_column=COL_NR)
    ws.cell(row=header_start_row, column=COL_NR, value="NR").font = font_bold_9; ws.cell(row=header_start_row, column=COL_NR).alignment = center_align

    # Formatif/Tugas & Sumatif Lingkup Materi (Merged across 1 row)
    ws.merge_cells(start_row=header_start_row, start_column=COL_FORMATIF_START, end_row=header_start_row, end_column=COL_FORMATIF_END)
    ws.cell(row=header_start_row, column=COL_FORMATIF_START, value="Formatif / Tugas").font = font_bold_9; ws.cell(row=header_start_row, column=COL_FORMATIF_START).alignment = center_align
    ws.merge_cells(start_row=header_start_row, start_column=COL_SUMATIF_START, end_row=header_start_row, end_column=COL_SUMATIF_END)
    ws.cell(row=header_start_row, column=COL_SUMATIF_START, value="Sumatif Lingkup Materi").font = font_bold_9; ws.cell(row=header_start_row, column=COL_SUMATIF_START).alignment = center_align

    # Row 2 & 3 Headers (TP and LM)
    for i in range(5):
        col_idx_f = COL_FORMATIF_START + i
        ws.cell(row=header_start_row + 1, column=col_idx_f, value=f"TP{i+1}").font = font_normal_9; ws.cell(row=header_start_row + 1, column=col_idx_f).alignment = center_align
        ws.merge_cells(start_row=header_start_row + 1, start_column=col_idx_f, end_row=header_start_row + 2, end_column=col_idx_f)

        col_idx_s = COL_SUMATIF_START + i
        ws.cell(row=header_start_row + 1, column=col_idx_s, value=f"LM{i+1}").font = font_normal_9; ws.cell(row=header_start_row + 1, column=col_idx_s).alignment = center_align
        ws.merge_cells(start_row=header_start_row + 1, start_column=col_idx_s, end_row=header_start_row + 2, end_column=col_idx_s)


    # Apply borders and set column width
    for r_idx in range(header_start_row, header_start_row + 3):
        for c_idx in range(1, TOTAL_EFFECTIVE_COLS_NILAI + 1):
            ws.cell(row=r_idx, column=c_idx).border = thin_border

    ws.row_dimensions[header_start_row].height = 20; ws.row_dimensions[header_start_row + 1].height = 20; ws.row_dimensions[header_start_row + 2].height = 20
    ws.column_dimensions[get_column_letter(COL_NO_START)].width = 2.57; ws.column_dimensions[get_column_letter(COL_NO_END)].width = 2.14
    ws.column_dimensions[get_column_letter(COL_NIS_START)].width = 2; ws.column_dimensions[get_column_letter(COL_NIS_START + 1)].width = 2; ws.column_dimensions[get_column_letter(COL_NIS_END)].width = 2
    ws.column_dimensions[get_column_letter(COL_NAMA_START)].width = 8.43; ws.column_dimensions[get_column_letter(COL_NAMA_END)].width = 23
    for i in range(COL_FORMATIF_START, COL_FORMATIF_END + 1): ws.column_dimensions[get_column_letter(i)].width = 3
    for i in range(COL_SUMATIF_START, COL_SUMATIF_END + 1): ws.column_dimensions[get_column_letter(i)].width = 3
    ws.column_dimensions[get_column_letter(COL_PTS)].width = 4; ws.column_dimensions[get_column_letter(COL_SAS_SAT)].width = 5; ws.column_dimensions[get_column_letter(COL_NR)].width = 4

    current_row += 3

    # --- Isi Data Siswa (dengan border di semua kolom) ---
    for i, row_data in enumerate(dataframe.itertuples(), 1):
        data_row_idx = current_row + i - 1

        # Merge cells for No., NIS, Nama
        ws.merge_cells(start_row=data_row_idx, start_column=COL_NO_START, end_row=data_row_idx, end_column=COL_NO_END)
        ws.merge_cells(start_row=data_row_idx, start_column=COL_NIS_START, end_row=data_row_idx, end_column=COL_NIS_END)
        ws.merge_cells(start_row=data_row_idx, start_column=COL_NAMA_START, end_row=data_row_idx, end_column=COL_NAMA_END)

        ws.cell(row=data_row_idx, column=COL_NO_START, value=i).font = font_normal_12
        ws.cell(row=data_row_idx, column=COL_NO_START).alignment = center_align

        # Menggunakan kolom 'NIS' dan 'Nama' dari DataFrame yang sudah difilter
        ws.cell(row=data_row_idx, column=COL_NIS_START, value=str(getattr(row_data, "NIS"))).font = font_normal_12
        ws.cell(row=data_row_idx, column=COL_NIS_START).alignment = center_align

        ws.cell(row=data_row_idx, column=COL_NAMA_START, value=getattr(row_data, "Nama")).font = font_normal_12
        ws.cell(row=data_row_idx, column=COL_NAMA_START).alignment = left_align

        for c_idx in range(1, TOTAL_EFFECTIVE_COLS_NILAI + 1):
            cell = ws.cell(row=data_row_idx, column=c_idx)
            cell.border = thin_border
            if c_idx >= COL_FORMATIF_START:
                cell.alignment = center_align

        ws.row_dimensions[data_row_idx].height = 20

    current_row += len(dataframe) + 1 # Pindah ke bawah tabel, 1 baris spasi

    # --- Bagian Tanda Tangan ---

    # 1. Mengetahui (Kepala Sekolah) di kolom C (start_column=3)
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=7)
    ws.cell(row=current_row, column=3, value="Mengetahui").font = font_normal_12; ws.cell(row=current_row, column=3).alignment = left_align

    # Guru Mapel - Tanggal (di kolom L/12)
    tz_jakarta = pytz.timezone('Asia/Jakarta')
    now = datetime.now(tz_jakarta)
    tanggal_formatted = f"Bantul, ............................... {now.year}"
    ws.merge_cells(start_row=current_row, start_column=12, end_row=current_row, end_column=20)
    # PERBAIKAN: Menulis nilai di sel awal (kolom 12), bukan kolom 14
    ws.cell(row=current_row, column=12, value=tanggal_formatted).font = font_normal_12; ws.cell(row=current_row, column=12).alignment = left_align

    current_row += 1

    # 2. Kepala Sekolah di kolom C (start_column=3)
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=7)
    ws.cell(row=current_row, column=3, value="Kepala Sekolah").font = font_normal_12; ws.cell(row=current_row, column=3).alignment = left_align

    # Guru Mapel (di kolom L/12)
    ws.merge_cells(start_row=current_row, start_column=12, end_row=current_row, end_column=20)
    # PERBAIKAN: Menulis nilai di sel awal (kolom 12), bukan kolom 14
    ws.cell(row=current_row, column=12, value=f"Guru {mapel}").font = font_normal_12; ws.cell(row=current_row, column=12).alignment = left_align

    current_row += 4 # Jarak untuk tanda tangan

    # 3. Nama Kepala Sekolah di kolom C (start_column=3)
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=7)
    ws.cell(row=current_row, column=3, value=KEPALA_SEKOLAH).font = font_normal_12; ws.cell(row=current_row, column=3).alignment = left_align

    # Nama Guru (di kolom L/12)
    ws.merge_cells(start_row=current_row, start_column=12, end_row=current_row, end_column=20)
    # PERBAIKAN: Menulis nilai di sel awal (kolom 12), bukan kolom 14
    ws.cell(row=current_row, column=12, value=guru).font = font_normal_12; ws.cell(row=current_row, column=12).alignment = left_align

    current_row += 1

    # 4. NIP Kepala Sekolah di kolom C (start_column=3)
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=7)
    ws.cell(row=current_row, column=3, value=f"NIP. {NIP_KEPSEK}").font = font_normal_12; ws.cell(row=current_row, column=3).alignment = left_align

    # NIP Guru (di kolom L/12)
    ws.merge_cells(start_row=current_row, start_column=12, end_row=current_row, end_column=20)
    # PERBAIKAN: Menulis nilai di sel awal (kolom 12), bukan kolom 14
    ws.cell(row=current_row, column=12, value=f"NIP. {nip_guru}").font = font_normal_12; ws.cell(row=current_row, column=12).alignment = left_align

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =========================================================
# Aplikasi Streamlit
# =========================================================
st.set_page_config(layout="wide")
st.title("🖨️ Form Cetak Daftar Nilai Siswa")

# --- Opsi Kelas dan Tahun Pelajaran ---
CLASS_OPTIONS = generate_class_options()
YEAR_OPTIONS = generate_academic_year_options(start_year=2025)

st.write("---")
st.header("⚙️ Pengaturan Form Daftar Nilai")

# --- Form Input Data Guru dan Kelas (Menggunakan Selectbox) ---
with st.container():
    col1, col2 = st.columns(2)

    with col1:
        mapel_terpilih = st.selectbox("Mata Pelajaran:", COMMON_SUBJECTS)
        # kelas_input sekarang akan berupa '7A', '8B', dst.
        kelas_input = st.selectbox("Kelas:", CLASS_OPTIONS, index=0)
        semester_input = st.selectbox("Semester:", ["Ganjil", "Genap"])

    with col2:
        tahun_pelajaran_input = st.selectbox("Tahun Pelajaran:", YEAR_OPTIONS)
        guru_input = st.text_input("Nama Guru Mata Pelajaran:", "")
        nip_guru_input = st.text_input("NIP Guru Mata Pelajaran:", "")

st.write("---")
st.header("🧑‍🎓 Data Siswa yang Diproses")

# --- Pemuatan Data dari CSV Default / Fallback ---
df_siswa_all = None
df_siswa_filtered = None

# Coba muat dari file CSV default
if os.path.exists(default_csv_path):
    try:
        df_siswa_all = pd.read_csv(default_csv_path)
        if 'Nama' in df_siswa_all.columns and 'NIS' in df_siswa_all.columns and 'Kelas' in df_siswa_all.columns:
            # Mengubah semua data menjadi string
            df_siswa_all = df_siswa_all[['NIS', 'Nama', 'Kelas']].astype(str)
            #st.success(f"Berhasil memuat {len(df_siswa_all)} data siswa dari **{default_csv_path}** (Total).")
        else:
            st.error(f"⚠️ **{default_csv_path}** harus memiliki kolom 'Nama', 'NIS', **dan 'Kelas'**.")
            df_siswa_all = None
    except Exception as e:
        st.error(f"❌ Gagal memuat file database default **{default_csv_path}**: {e}")
        df_siswa_all = None

# Fallback ke data hardcoded jika CSV default gagal dimuat atau tidak ada
if df_siswa_all is None or df_siswa_all.empty:
    data_default = {
        'NIS': ['1001', '1002', '1003', '1004', '1005', '1006'],
        'Nama': ['Budi Santoso', 'Siti Aminah', 'Joko Susanto', 'Dewi Puspita', 'Rizky Pratama', 'Aulia Putri'],
        # DATA FALLBACK SEKARANG JUGA TANPA SPASI
        'Kelas': ['7A', '7A', '8B', '9C', '7B', '7A'],
    }
    df_siswa_all = pd.DataFrame(data_default)
    st.warning(f"File database default **{default_csv_path}** tidak ditemukan atau error. Menggunakan contoh data default (format: **7A**).")


# --- LOGIKA FILTERING BERDASARKAN KELAS PILIHAN ---
if df_siswa_all is not None and not df_siswa_all.empty:
    # Filter DataFrame: df_siswa_all['Kelas'] (misal: '7A') == kelas_input (misal: '7A')
    df_siswa_filtered = df_siswa_all[df_siswa_all['Kelas'] == kelas_input].reset_index(drop=True)

    # Tampilkan data yang sudah difilter
    if len(df_siswa_filtered) > 0:
        st.subheader(f"Daftar Siswa Kelas {kelas_input}")
        st.dataframe(df_siswa_filtered[['NIS', 'Nama']], use_container_width=True)
        st.success(f"✅ Siap mencetak **{len(df_siswa_filtered)}** siswa untuk Kelas **{kelas_input}**.")
    else:
        st.warning(f"⚠️ **Tidak ditemukan** data siswa untuk Kelas **{kelas_input}** di database. Pastikan penulisan kelas di file CSV Anda adalah **Angka** diikuti **Huruf Kelas tanpa spasi** (contoh: '7A').")
        df_siswa_filtered = None
else:
    df_siswa_filtered = None


st.write("---")

# --- Tombol Cetak Form Nilai Siswa ---
if df_siswa_filtered is not None and len(df_siswa_filtered) > 0:
    st.header("Aksi Cetak")

    # Tombol Cetak Form Nilai Siswa
    excel_buffer_nilai = generate_excel_form_nilai_siswa(
        df_siswa_filtered,
        mapel_terpilih,
        semester_input,
        kelas_input,
        tahun_pelajaran_input,
        guru_input,
        nip_guru_input
    )
    st.download_button(
        label="⬇️ Cetak Form Nilai Siswa (Excel)",
        # File name juga menggunakan format tanpa spasi
        file_name=f"Form_Nilai_Siswa_{kelas_input}_{mapel_terpilih}_{semester_input}.xlsx",
        data=excel_buffer_nilai,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_nilai"
    )
    st.success("Form Nilai Siswa siap dicetak!")
else:
    st.error("❌ **Tidak ada data siswa yang valid untuk dicetak.** Silakan pilih kelas lain atau pastikan kolom 'Kelas' di file `daftar_siswa.csv` Anda menggunakan format Angka tanpa Spasi (misal: **'7A'**).")
